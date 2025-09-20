import os
import io
import time
from datetime import datetime

import streamlit as st
import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from zipfile import BadZipFile

import requests
import nfl_data_py as nfl
from urllib.error import HTTPError, URLError

# --- App setup ---
st.set_page_config(page_title="Bears Weekly Tracker", layout="wide")

DATA_DIR = "./data"
EXPORTS_DIR = "./exports"
CSV_DIR = "./csv"
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(EXPORTS_DIR, exist_ok=True)
os.makedirs(CSV_DIR, exist_ok=True)

EXCEL_PATH = os.path.join(DATA_DIR, "bears_weekly_analytics.xlsx")

# Sheet names
SHEET_OFFENSE = "Offense"
SHEET_DEFENSE = "Defense"
SHEET_PERSONNEL = "Personnel"
SHEET_SNAP_COUNTS = "SnapCounts"
SHEET_INJURIES = "Injuries"
SHEET_MEDIA = "MediaSummaries"
SHEET_OPP_PREVIEW = "OpponentPreview"
SHEET_PREDICTIONS = "Predictions"
SHEET_NFL_AVG_MANUAL = "NFL_Averages_Manual"
SHEET_YTD_TEAM_OFF = "YTD_Team_Offense"
SHEET_YTD_TEAM_DEF = "YTD_Team_Defense"
SHEET_YTD_NFL_OFF = "YTD_NFL_Offense"
SHEET_YTD_NFL_DEF = "YTD_NFL_Defense"
SHEET_NFL_WEEKLY_AVG = "NFL_Weekly_Averages"

ALL_SHEETS = [
    SHEET_OFFENSE, SHEET_DEFENSE, SHEET_PERSONNEL, SHEET_SNAP_COUNTS,
    SHEET_INJURIES, SHEET_MEDIA, SHEET_OPP_PREVIEW, SHEET_PREDICTIONS,
    SHEET_NFL_AVG_MANUAL, SHEET_YTD_TEAM_OFF, SHEET_YTD_TEAM_DEF,
    SHEET_YTD_NFL_OFF, SHEET_YTD_NFL_DEF, SHEET_NFL_WEEKLY_AVG
]

NFL_TEAM = "CHI"
DEFAULT_SEASON = 2025
DEFAULT_WEEK = 2

# ---------------------------
# Safe import wrapper
# ---------------------------
def import_df_safely(import_fn, seasons):
    """
    Call a data importer and return a DataFrame.
    Any exception returns an empty DataFrame and logs a short, friendly note.
    """
    try:
        return import_fn(seasons)
    except HTTPError as e:
        code = getattr(e, "code", "HTTPError")
        if code == 404:
            st.info(f"{import_fn.__name__} for {seasons}: 404 (not posted yet).")
        else:
            st.info(f"{import_fn.__name__} for {seasons}: HTTP {code}.")
    except URLError:
        st.info(f"{import_fn.__name__} network error.")
    except NameError:
        # e.g., NameError("name 'Error' is not defined") seen on early-season pbp
        st.info(f"{import_fn.__name__} is not ready for {seasons} yet.")
    except Exception:
        st.info(f"{import_fn.__name__} not available for {seasons}.")
    return pd.DataFrame()

# ---------------------------
# Excel helpers
# ---------------------------
def ensure_xlsx(path):
    if not path.lower().endswith(".xlsx"):
        base, _ = os.path.splitext(path)
        path = base + ".xlsx"
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_OFFENSE
        for name in ALL_SHEETS:
            if name != SHEET_OFFENSE:
                wb.create_sheet(title=name)
        wb.save(path)
        return path
    try:
        _ = load_workbook(path, read_only=True)
    except (BadZipFile, KeyError, OSError):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_OFFENSE
        for name in ALL_SHEETS:
            if name != SHEET_OFFENSE:
                wb.create_sheet(title=name)
        wb.save(path)
    return path

def read_sheet(path, sheet):
    ensure_xlsx(path)
    try:
        df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
        df.columns = [str(c).strip() for c in df.columns]
        return df
    except Exception:
        return pd.DataFrame()

def write_sheet(path, sheet, df):
    ensure_xlsx(path)
    try:
        with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as xlw:
            df.to_excel(xlw, index=False, sheet_name=sheet)
    except FileNotFoundError:
        with pd.ExcelWriter(path, engine="openpyxl", mode="w") as xlw:
            df.to_excel(xlw, index=False, sheet_name=sheet)

def append_to_sheet(path, sheet, df_new, dedupe_on=None):
    df_old = read_sheet(path, sheet)
    if df_old.empty:
        df_out = df_new.copy()
    else:
        df_out = pd.concat([df_old, df_new], ignore_index=True)
    if dedupe_on:
        df_out = df_out.drop_duplicates(subset=dedupe_on, keep="last")
    write_sheet(path, sheet, df_out)
    return df_out

# ---------------------------
# CSV helpers
# ---------------------------
def csv_name(base, season, week, team):
    return os.path.join(CSV_DIR, f"{base}_{int(season)}_W{int(week):02d}_{team}.csv")

def save_csv(df, base, season, week, team):
    path = csv_name(base, season, week, team)
    os.makedirs(CSV_DIR, exist_ok=True)
    df.to_csv(path, index=False)
    return path

# ---------------------------
# Small utils (NEW: de-dup columns + numeric check)
# ---------------------------
def _normalize_team_col(df):
    if df.empty:
        return df
    if 'team' in df.columns:
        return df
    if 'recent_team' in df.columns:
        df = df.rename(columns={'recent_team': 'team'})
    if 'posteam' in df.columns:
        df = df.rename(columns={'posteam': 'team'})
    return df

def select_numeric(df):
    return [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]

def make_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure DataFrame has unique column names for Arrow/Streamlit."""
    if df is None or df.empty:
        return df
    cols = list(df.columns)
    seen = {}
    new_cols = []
    for c in cols:
        if c not in seen:
            seen[c] = 1
            new_cols.append(c)
        else:
            k = seen[c]
            seen[c] = k + 1
            new_cols.append(f"{c}.{k}")  # append a numeric suffix
    df = df.copy()
    df.columns = new_cols
    return df

def has_meaningful_numeric(df: pd.DataFrame) -> bool:
    """True if the df has any numeric column with at least one non-null value."""
    if df is None or df.empty:
        return False
    numeric_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    if not numeric_cols:
        return False
    return pd.notna(df[numeric_cols]).any().any()

# ---------------------------
# GitHub Release asset discovery (avoids hard-coded filenames)
# ---------------------------
GITHUB_API = "https://api.github.com/repos/nflverse/nflverse-data/releases/tags/{tag}"
GITHUB_DL  = "https://github.com/nflverse/nflverse-data/releases/download/{tag}/{name}"

def _gh_get_tag_assets(tag, token=None):
    """
    Return list of assets for a release tag (e.g., 'stats_team', 'player_stats').
    If you set a GITHUB_TOKEN env var, you get higher rate limits.
    """
    url = GITHUB_API.format(tag=tag)
    headers = {"Accept": "application/vnd.github+json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    try:
        r = requests.get(url, headers=headers, timeout=20)
        if r.status_code != 200:
            return []
        data = r.json()
        return data.get("assets", []) or []
    except Exception:
        return []

def _find_asset_url(tag, want_year, contains):
    """
    Find an asset whose name contains BOTH the target year and the substring `contains` (case-insensitive).
    Returns the direct download URL string if found, else None.
    """
    token = os.getenv("GITHUB_TOKEN")
    assets = _gh_get_tag_assets(tag, token=token)
    if not assets:
        return None

    year_str = str(want_year)
    substr = contains.lower()
    for a in assets:
        name = a.get("name", "")
        if not name:
            continue
        if year_str in name and substr in name.lower():
            return GITHUB_DL.format(tag=tag, name=name)
    return None

def get_nflverse_url_team_week(year):
    # e.g., stats_team_week_2025.csv or stats_team_week_2025_v2.csv
    return _find_asset_url(tag="stats_team", want_year=year, contains="stats_team_week")

def get_nflverse_url_player_week(year):
    # e.g., stats_player_week_2025.csv
    return _find_asset_url(tag="player_stats", want_year=year, contains="stats_player_week")

def _read_csv_with_retries(url, retries=2, sleep_sec=1.0):
    last_err = None
    for _ in range(retries + 1):
        try:
            # cache-buster to dodge stale CDN caches
            u = f"{url}?t={int(time.time())}"
            return pd.read_csv(u)
        except Exception as e:
            last_err = e
            time.sleep(sleep_sec)
    raise last_err if last_err else RuntimeError("Failed to fetch CSV")

# ---------------------------
# Direct weekly loaders via nflverse releases (with discovery)
# ---------------------------
def load_team_week_from_release(season: int) -> pd.DataFrame:
    url = get_nflverse_url_team_week(season)
    if not url:
        raise RuntimeError("team-week asset not found in release tag 'stats_team'")
    df = _read_csv_with_retries(url)
    return df

def load_team_week_via_players(season: int) -> pd.DataFrame:
    url = get_nflverse_url_player_week(season)
    if not url:
        raise RuntimeError("player-week asset not found in release tag 'player_stats'")
    dfp = _read_csv_with_retries(url)

    group_cols = [c for c in ["season", "week", "team"] if c in dfp.columns]
    if not all(g in dfp.columns for g in group_cols):
        return pd.DataFrame()

    num_cols = [c for c in dfp.columns if pd.api.types.is_numeric_dtype(dfp[c])]
    # aggregate player-week → team-week
    df_team = dfp.groupby(group_cols, as_index=False)[num_cols].sum()
    return df_team

# ---------------------------
# ESPN fallback (improved) – used only if nflverse week CSVs not posted yet
# ---------------------------
def _espn_get(url, params=None):
    try:
        r = requests.get(url, params=params, timeout=20)
        r.raise_for_status()
        return r.json()
    except Exception:
        return None

def espn_week_events(season: int, week: int):
    """
    ESPN scoreboard for regular season (seasontype=2).
    """
    url = "https://site.api.espn.com/apis/site/v2/sports/football/nfl/scoreboard"
    data = _espn_get(url, params={"year": season, "week": week, "seasontype": 2})
    if not data or "events" not in data:
        return []
    return data["events"]

def espn_game_summary(event_id: str):
    """
    ESPN summary (includes team boxscore).
    """
    url = "https://site.api.espn.com/apis/site/v2/sports/football/nfl/summary"
    return _espn_get(url, params={"event": event_id}) or {}

def _abbr_from_competitor(comp):
    try:
        return comp["team"]["abbreviation"]
    except Exception:
        return None

def _points_from_competitor(comp):
    try:
        return float(comp.get("score"))
    except Exception:
        return None

def _to_number(x):
    """
    Convert ESPN values (which may be strings in displayValue) into floats.
    - "30:12" -> minutes (30 + 12/60) for time of possession
    - "54.2%" -> 0.542
    """
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s == "":
        return None
    # Time of possession "MM:SS"
    if ":" in s and s.count(":") == 1:
        mm, ss = s.split(":")
        try:
            return float(mm) + float(ss)/60.0
        except Exception:
            return None
    # Percent "54.2%"
    if s.endswith("%"):
        try:
            return float(s[:-1]) / 100.0
        except Exception:
            return None
    # Simple number
    try:
        return float(s.replace(",", ""))
    except Exception:
        return None

def _team_stats_from_summary(summary, team_abbr):
    """
    Parse team-level stats from ESPN summary JSON → dict for team_abbr.
    Reads both 'value' and 'displayValue'. Normalizes names and converts to floats where possible.
    Handles ratios like "6-12" (third downs) into Made/Att/Rate.
    """
    out = {}
    try:
        box = summary.get("boxscore", {})
        teams = box.get("teams", [])
        for t in teams:
            abbr = t.get("team", {}).get("abbreviation")
            if abbr != team_abbr:
                continue
            # points (also available from competition)
            if "score" in t:
                out["points"] = _to_number(t["score"])
            stats = t.get("statistics", [])
            for s in stats:
                # ESPN commonly provides: name, displayName, value, displayValue
                name = s.get("name") or s.get("displayName")
                if not name:
                    continue
                # Prefer numeric value; if missing, try displayValue
                val = s.get("value", None)
                if val is None:
                    val = s.get("displayValue", None)

                # Normalize key names we care about
                key = (
                    name.replace(" ", "_")
                        .replace("%", "Pct")
                        .replace("/", "_per_")
                        .replace("-", "_")
                        .strip()
                )

                # A few ESPN alternates → canonical
                aliases = {
                    "avgYardsPerPlay": "yardsPerPlay",
                    "totalYardsGained": "totalYards",
                    "rushYards": "rushingYards",
                    "rushAttempts": "rushingAttempts",
                    "passYards": "passingYards",
                    "possessionTime": "timeOfPossession",
                    "avgGain": "yardsPerPlay",
                    "thirdDowns": "thirdDowns",   # may come as ratio text
                    "redZone": "redZone",         # may come as ratio text
                }
                key = aliases.get(key, key)

                # Convert to numeric where sensible
                if key in ("timeOfPossession",):
                    num = _to_number(val)  # -> minutes as float
                else:
                    num = _to_number(val)

                # Ratios like "6-12" or "6/12" (third downs, red zone)
                if num is None and isinstance(val, str) and any(sep in val for sep in ("-", "/")):
                    s2 = val.replace("-", "/")
                    parts = [p.strip() for p in s2.split("/") if p.strip() != ""]
                    if len(parts) == 2:
                        try:
                            made, att = float(parts[0]), float(parts[1])
                            base = key
                            out[base + "Made"] = made
                            out[base + "Att"] = att
                            if att > 0:
                                out[base + "Rate"] = made / att
                            continue
                        except Exception:
                            pass

                # Store numeric if we have it
                if num is not None:
                    out[key] = num

    except Exception:
        return {}
    return out

def espn_team_week_for_week(season: int, week: int) -> pd.DataFrame:
    events = espn_week_events(season, week)
    # quick diagnostic to show if ESPN has anything for the week
    st.caption(f"ESPN events for {season} W{week}: {len(events)}")

    if not events:
        return pd.DataFrame()

    rows = []
    for ev in events:
        event_id = ev.get("id")
        competitions = ev.get("competitions", [])
        if not competitions:
            continue
        comp = competitions[0]
        competitors = comp.get("competitors", [])
        pts_by_abbr = {}
        for c in competitors:
            ab = _abbr_from_competitor(c)
            pts_by_abbr[ab] = _points_from_competitor(c)

        summary = espn_game_summary(event_id)
        for c in competitors:
            ab = _abbr_from_competitor(c)
            if not ab:
                continue
            stats = _team_stats_from_summary(summary, ab)

            row = {"season": season, "week": week, "team": ab}
            # Preferred keys we try to populate
            wanted = [
                "points",
                "totalYards",
                "yardsPerPlay",
                "passingYards",
                "completions",
                "attempts",
                "completionPercentage",
                "rushingYards",
                "rushingAttempts",
                "sacks",
                "turnovers",
                "firstDowns",
                # derived/alternate names
                "thirdDownsRate", "thirdDownsMade", "thirdDownsAtt",
                "redZoneRate", "redZoneScores", "redZoneAttempts",
                "penalties", "penaltyYards",
                "timeOfPossession",
            ]

            # points first: prefer parsed, else from competition
            row["points"] = stats.get("points", pts_by_abbr.get(ab))

            # fill other wanted if present
            for k in wanted:
                if k == "points":
                    continue
                if k in stats:
                    row[k] = stats[k]

            rows.append(row)

    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    # Coerce numerics
    for c in df.columns:
        if c in ("season", "week", "team"):
            continue
        df[c] = pd.to_numeric(df[c], errors="coerce")
    return df

# ---------------------------
# PBP fallback aggregation (last resort)
# ---------------------------
def _pbp_week_for_season(season: int) -> pd.DataFrame:
    try:
        pbp = import_df_safely(nfl.import_pbp_data, [season])
        if pbp.empty:
            return pd.DataFrame()
        if 'season_type' in pbp.columns:
            pbp = pbp[pbp['season_type'].astype(str).str.upper().str.contains("REG")]
        return pbp
    except Exception:
        return pd.DataFrame()

def _aggregate_team_week_from_pbp(pbp: pd.DataFrame, team: str, week: int) -> pd.DataFrame:
    if pbp.empty:
        return pd.DataFrame()
    if 'week' in pbp.columns:
        pbp = pbp[pbp['week'] == week]
    off = pbp[pbp.get('posteam') == team].copy()
    if off.empty:
        return pd.DataFrame()
    off['is_pass'] = off.get('pass', 0).fillna(0).astype(int)
    off['is_rush'] = off.get('rush', 0).fillna(0).astype(int)
    row = {
        "team": team,
        "week": week,
        "plays": len(off),
        "yards_gained": float(off.get('yards_gained', 0).fillna(0).sum()),
        "pass_att": int(off.get('pass_attempt', 0).fillna(0).sum()),
        "completions": int(off.get('complete_pass', 0).fillna(0).sum()),
        "pass_yards": float(off.loc[off['is_pass'] == 1, 'yards_gained'].fillna(0).sum()),
        "rush_att": int(off.get('rush_attempt', 0).fillna(0).sum()),
        "rush_yards": float(off.loc[off['is_rush'] == 1, 'yards_gained'].fillna(0).sum()),
        "sacks": int(off.get('sack', 0).fillna(0).sum()),
        "interceptions": int(off.get('interception', 0).fillna(0).sum()),
        "fumbles_lost": int(off.get('fumble_lost', 0).fillna(0).sum()),
    }
    return pd.DataFrame([row])

# ---------------------------
# Unified weekly resolver (Team-week → Player-week agg → nfl_data_py → ESPN(week) → empty)
# ---------------------------
def resolve_team_week_table(season: int, prefer_week: int = None) -> pd.DataFrame:
    # 1) team-week release
    try:
        df = load_team_week_from_release(season)
        if not df.empty:
            return df
    except Exception as e:
        st.caption(f"Team-week release not available yet: {e}")

    # 2) player-week aggregate
    try:
        df = load_team_week_via_players(season)
        if not df.empty:
            return df
    except Exception as e:
        st.caption(f"Player-week release not available yet: {e}")

    # 3) nfl_data_py weekly (works for past seasons)
    weekly = import_df_safely(nfl.import_weekly_data, [season])
    weekly = _normalize_team_col(weekly)
    if not weekly.empty:
        return weekly

    # 4) ESPN fallback for the specific week
    if prefer_week is not None:
        espn_df = espn_team_week_for_week(season, prefer_week)
        if not espn_df.empty:
            return espn_df

    # 5) give up → caller may still use PBP per-team
    return pd.DataFrame()

# ---------------------------
# Convenience selectors
# ---------------------------
def bears_row_for_week(team_week: pd.DataFrame, season: int, week: int, team: str) -> pd.DataFrame:
    if team_week.empty:
        return pd.DataFrame()
    want = team_week.copy()
    for col, val in (("season", season), ("week", week), ("team", team)):
        if col in want.columns:
            want = want[want[col] == val]
    return want.copy()

def nfl_week_average_from_table(team_week: pd.DataFrame, week: int) -> pd.DataFrame:
    if team_week.empty or "week" not in team_week.columns:
        return pd.DataFrame()
    wk = team_week[team_week["week"] == week]
    if wk.empty:
        return pd.DataFrame()
    nums = select_numeric(wk)
    if not nums:
        return pd.DataFrame()
    out = wk[nums].mean(numeric_only=True).to_frame().T
    out.columns = [f"NFL_Avg._{c}" for c in out.columns]
    return out.reset_index(drop=True)

def fetch_nfl_averages_YTD(season, upto_week, weekly_table=None):
    if weekly_table is None:
        weekly_table = resolve_team_week_table(season, prefer_week=upto_week)
    if not weekly_table.empty and 'week' in weekly_table.columns:
        df = weekly_table[(weekly_table['week'] >= 1) & (weekly_table['week'] <= upto_week)].copy()
        if not df.empty:
            nums = select_numeric(df)
            if not nums:
                return pd.DataFrame()
            out = df[nums].mean(numeric_only=True).to_frame().T
            out.columns = [f"NFL_Avg._{c}" for c in out.columns]
            return out.reset_index(drop=True)
    return pd.DataFrame()

def fetch_bears_averages_YTD(season, upto_week, team, weekly_table=None):
    if weekly_table is None:
        weekly_table = resolve_team_week_table(season, prefer_week=upto_week)
    if not weekly_table.empty and 'week' in weekly_table.columns:
        chi = weekly_table[(weekly_table['team'] == team) &
                           (weekly_table['week'] >= 1) &
                           (weekly_table['week'] <= upto_week)].copy()
        if not chi.empty:
            nums = select_numeric(chi)
            if not nums:
                return pd.DataFrame()
            return chi[nums].mean(numeric_only=True).to_frame().T.reset_index(drop=True)
    return pd.DataFrame()

# ---------------------------
# Snap counts & injuries
# ---------------------------
def fetch_snap_counts(season, team):
    snaps = import_df_safely(nfl.import_snap_counts, [season])
    if snaps.empty:
        return pd.DataFrame()
    snaps = _normalize_team_col(snaps)
    return snaps[snaps["team"] == team].copy()

def fetch_snap_counts_week_both(season, week, team, opp):
    snaps = import_df_safely(nfl.import_snap_counts, [season])
    if snaps.empty:
        return pd.DataFrame(), pd.DataFrame()
    snaps = _normalize_team_col(snaps)
    has_week = 'week' in snaps.columns
    chi = snaps[(snaps["team"] == team) & ((snaps["week"] == week) if has_week else True)].copy()
    oppdf = snaps[(snaps["team"] == opp) & ((snaps["week"] == week) if has_week else True)].copy()
    return chi, oppdf

def fetch_injuries_for_week(season: int, week: int, team: str) -> pd.DataFrame:
    inj = import_df_safely(nfl.import_injuries, [season])
    if inj.empty:
        return pd.DataFrame()
    inj = _normalize_team_col(inj)
    if 'week' in inj.columns:
        return inj[(inj['team'] == team) & (inj['week'] == week)].copy()
    return inj[inj['team'] == team].copy()

# ---------------------------
# PDF export (optional)
# ---------------------------
try:
    from fpdf import FPDF
except Exception:
    FPDF = None

def export_week_pdf(season, week, opponent, summary_lines):
    if FPDF is None:
        return b""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=14)
    pdf.cell(0, 10, f"Bears Weekly Report — {season} Week {week} vs {opponent}", ln=1)
    pdf.set_font("Arial", size=11)
    for line in summary_lines:
        pdf.multi_cell(0, 8, txt=line)
    out = io.BytesIO()
    pdf.output(out, "F")
    return out.getvalue()

HEADLINE_KEYS = ["points","yards","pass","rush","sack","int","fumble","third","red","rz"]

def _first_n_common_numeric(chi, opp, n=6):
    if chi.empty or opp.empty:
        return []
    nums_chi = [c for c in chi.columns if pd.api.types.is_numeric_dtype(chi[c])]
    nums_opp = [c for c in opp.columns if pd.api.types.is_numeric_dtype(opp[c])]
    common = [c for c in nums_chi if c in nums_opp]
    def score(col):
        lc = col.lower()
        return max((1 if kw in lc else 0) for kw in HEADLINE_KEYS)
    common.sort(key=lambda c: (-score(c), c.lower()))
    return common[:n]

def _safe_get_scalar(df, col):
    if df.empty or col not in df.columns:
        return None
    try:
        v = float(df[col].iloc[0])
        if np.isfinite(v):
            return v
    except Exception:
        pass
    return None

def _fmt(v):
    return "-" if v is None else (f"{v:.2f}" if abs(v) < 1000 else f"{v:,.0f}")

def build_pdf_summary_lines(season, week, opponent, weekly_table):
    lines = []
    chi = bears_row_for_week(weekly_table, season, week, NFL_TEAM)
    opp = bears_row_for_week(weekly_table, season, week, opponent)

    if chi.empty and opp.empty:
        # As a last-ditch, try PBP per-team to populate something
        pbp = _pbp_week_for_season(season)
        chi = _aggregate_team_week_from_pbp(pbp, NFL_TEAM, week)
        opp = _aggregate_team_week_from_pbp(pbp, opponent, week)

    nfl_wk = nfl_week_average_from_table(weekly_table, week)

    lines.append(f"Season {season} — Week {week} vs {opponent}")
    lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append("")
    lines.append("Bears vs Opponent — This Week")
    if chi.empty and opp.empty:
        lines.append("  • No week rows available.")
    else:
        for c in _first_n_common_numeric(chi, opp, n=6):
            chi_v = _safe_get_scalar(chi, c)
            opp_v = _safe_get_scalar(opp, c)
            diff = None if (chi_v is None or opp_v is None) else chi_v - opp_v
            lines.append(f"  • {c}: CHI {_fmt(chi_v)} | OPP {_fmt(opp_v)} | Δ {_fmt(diff)}")

    lines.append("")
    lines.append("Bears vs NFL Average — This Week")
    if not chi.empty and not nfl_wk.empty:
        nfl_map = {k.replace("NFL_Avg._", ""): k for k in nfl_wk.columns if k.startswith("NFL_Avg._")}
        nums_chi = [c for c in chi.columns if pd.api.types.is_numeric_dtype(chi[c])]
        both = [c for c in nums_chi if c in nfl_map]
        for c in both[:6]:
            chi_v = _safe_get_scalar(chi, c)
            nfl_v = _safe_get_scalar(nfl_wk, nfl_map[c])
            diff = None if (chi_v is None or nfl_v is None) else chi_v - nfl_v
            lines.append(f"  • {c}: CHI {_fmt(chi_v)} | NFLwk {_fmt(nfl_v)} | Δ {_fmt(diff)}")
    else:
        lines.append("  • NFL weekly average not available.")

    lines.append("")
    lines.append("Bears YTD vs NFL YTD — Weeks 1..W")
    bears_ytd = fetch_bears_averages_YTD(season, week, NFL_TEAM, weekly_table)
    nfl_ytd = fetch_nfl_averages_YTD(season, week, weekly_table)
    if bears_ytd.empty or nfl_ytd.empty:
        lines.append("  • YTD rows not available yet.")
    else:
        nfl_map_ytd = {k.replace("NFL_Avg._", ""): k for k in nfl_ytd.columns if k.startswith("NFL_Avg._")}
        nums_brs = [c for c in bears_ytd.columns if pd.api.types.is_numeric_dtype(bears_ytd[c])]
        both_ytd = [c for c in nums_brs if c in nfl_map_ytd]
        for c in both_ytd[:6]:
            brs_v = _safe_get_scalar(bears_ytd, c)
            nfl_v = _safe_get_scalar(nfl_ytd, nfl_map_ytd[c])
            diff = None if (brs_v is None or nfl_v is None) else brs_v - nfl_v
            lines.append(f"  • {c}: CHI_YTD {_fmt(brs_v)} | NFL_YTD {_fmt(nfl_v)} | Δ {_fmt(diff)}")
    lines.append("")
    return lines

# ---------------------------
# Sidebar
# ---------------------------
st.sidebar.title("Controls")
season_input = st.sidebar.number_input("Season", min_value=2012, max_value=2030, value=DEFAULT_SEASON, step=1)
week_input = st.sidebar.number_input("Week", min_value=1, max_value=22, value=DEFAULT_WEEK, step=1)
opponent_input = st.sidebar.text_input("Opponent (3-letter code)", value="MIN").strip().upper()

st.sidebar.markdown("---")
st.sidebar.subheader("NFL Updates (Direct nflverse → ESPN → Fallbacks)")
if st.sidebar.button("Fetch NFL Data (Auto)"):
    team_week = resolve_team_week_table(int(season_input), prefer_week=int(week_input))
    if team_week.empty:
        st.sidebar.error("No weekly table available yet (team-week/player-week/ESPN).")
    else:
        nfl_ytd = fetch_nfl_averages_YTD(int(season_input), int(week_input), team_week)
        if nfl_ytd.empty:
            st.sidebar.warning("Could not compute NFL YTD averages.")
        else:
            write_sheet(EXCEL_PATH, SHEET_YTD_NFL_OFF, nfl_ytd)
            write_sheet(EXCEL_PATH, SHEET_YTD_NFL_DEF, nfl_ytd)
            st.sidebar.success("NFL YTD averages saved.")

        wkavg = nfl_week_average_from_table(team_week, int(week_input))
        if not wkavg.empty:
            write_sheet(EXCEL_PATH, SHEET_NFL_WEEKLY_AVG, wkavg.assign(week=int(week_input)))
            st.sidebar.success("NFL weekly average (selected week) saved.")
        else:
            st.sidebar.info("No league weekly average for that week yet.")

st.sidebar.subheader("Snap Updates")
if st.sidebar.button("Fetch Snap Counts"):
    snaps = fetch_snap_counts(int(season_input), NFL_TEAM)
    if snaps.empty:
        st.sidebar.warning("No snap counts fetched.")
    else:
        keep = [c for c in snaps.columns if c in (
            "season", "week", "team", "player", "position",
            "offense_snaps", "defense_snaps", "special_teams_snaps",
            "offense_pct", "defense_pct", "special_teams_pct"
        )]
        if keep:
            snaps = snaps[keep]
        append_to_sheet(EXCEL_PATH, SHEET_SNAP_COUNTS, snaps,
                        dedupe_on=["season", "week", "team", "player"] if "player" in snaps.columns else ["season", "week", "team"])
        st.sidebar.success(f"Snap counts saved for {NFL_TEAM} {int(season_input)}.")

st.sidebar.subheader("Injuries")
if st.sidebar.button("Fetch Injuries (Auto)"):
    inj = fetch_injuries_for_week(int(season_input), int(week_input), NFL_TEAM)
    if inj.empty:
        st.sidebar.warning("No injury rows found for this week/team.")
    else:
        append_to_sheet(EXCEL_PATH, SHEET_INJURIES, inj,
                        dedupe_on=[c for c in ["season", "week", "team", "player", "report_date"] if c in inj.columns])
        st.sidebar.success(f"Injuries saved for {NFL_TEAM} W{int(week_input)}.")

st.sidebar.subheader("CSV Outputs")
if st.sidebar.button("Write Weekly CSVs (Auto)"):
    season_val, week_val = int(season_input), int(week_input)
    team_week = resolve_team_week_table(season_val, prefer_week=week_val)
    if team_week.empty:
        st.sidebar.error("No weekly table available yet to write CSVs.")
    else:
        chi = bears_row_for_week(team_week, season_val, week_val, NFL_TEAM)
        opp = bears_row_for_week(team_week, season_val, week_val, opponent_input)
        wrote = False
        if not chi.empty and has_meaningful_numeric(chi):
            p1 = save_csv(make_unique_columns(chi), "Offense", season_val, week_val, NFL_TEAM)
            p2 = save_csv(make_unique_columns(chi), "Defense", season_val, week_val, NFL_TEAM)
            st.sidebar.success(f"Wrote Bears weekly CSVs:\n• {os.path.basename(p1)}\n• {os.path.basename(p2)}")
            wrote = True
        if not opp.empty and has_meaningful_numeric(opp):
            p1 = save_csv(make_unique_columns(opp), "Offense", season_val, week_val, opponent_input)
            p2 = save_csv(make_unique_columns(opp), "Defense", season_val, week_val, opponent_input)
            st.sidebar.success(f"Wrote Opponent weekly CSVs:\n• {os.path.basename(p1)}\n• {os.path.basename(p2)}")
            wrote = True
        if not wrote:
            st.sidebar.info("No CHI/OPP rows with numeric values yet; try later.")

st.sidebar.markdown("---")
st.sidebar.subheader("Downloads")
if os.path.exists(EXCEL_PATH):
    with open(EXCEL_PATH, "rb") as f:
        st.sidebar.download_button(
            label="Download Excel (All Data)",
            data=f.read(),
            file_name="bears_weekly_analytics.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
else:
    st.sidebar.info("Excel not found yet; upload or fetch to create it.")

pdf_stub_path = os.path.join(EXPORTS_DIR, f"{int(season_input)}_W{int(week_input):02d}_{opponent_input}_Final.pdf")
if os.path.exists(pdf_stub_path):
    with open(pdf_stub_path, "rb") as f:
        st.sidebar.download_button(
            label=f"Download Final PDF ({int(season_input)} W{int(week_input):02d})",
            data=f.read(),
            file_name=os.path.basename(pdf_stub_path),
            mime="application/pdf"
        )
else:
    st.sidebar.caption("Final PDF will appear here after you export it.")

# ---------------------------
# Main
# ---------------------------
st.title("Chicago Bears Weekly Tracker")
st.markdown("**Order:** 1) Weekly Controls → 2) Upload Weekly Data → 3) NFL Averages → 4) Comparisons → 5) Notes → 6) Exports")

c1, c2, c3 = st.columns([1, 1, 2])
with c1:
    st.write(f"**Season:** {int(season_input)} | **Week:** {int(week_input)}")
with c2:
    st.write(f"**Opponent:** {opponent_input}")
with c3:
    st.caption("Use the sidebar to change Season, Week, and Opponent.")

# 1) Weekly Controls
with st.expander("1) Weekly Controls", expanded=True):
    st.info("Direct nflverse weekly → player-aggregate → nfl_data_py → ESPN (week) → PBP (last resort).")

    # Diagnostics expander to confirm discovery works
    with st.expander("Diagnostics — nflverse releases", expanded=False):
        y = int(season_input)
        if st.button("Check nflverse release assets now"):
            team_url = get_nflverse_url_team_week(y)
            player_url = get_nflverse_url_player_week(y)
            st.write("Team-week asset URL:", team_url or "⛔ not found in 'stats_team' assets")
            st.write("Player-week asset URL:", player_url or "⛔ not found in 'player_stats' assets")
            if team_url:
                st.write("HEAD check team-week…")
                try:
                    r = requests.head(team_url, timeout=15, allow_redirects=True)
                    st.write("status:", r.status_code)
                except Exception as e:
                    st.write("HEAD error:", e)
            if player_url:
                st.write("HEAD check player-week…")
                try:
                    r = requests.head(player_url, timeout=15, allow_redirects=True)
                    st.write("status:", r.status_code)
                except Exception as e:
                    st.write("HEAD error:", e)

    cc1, cc2 = st.columns(2)
    with cc1:
        st.markdown("**Quick Fetch (This Week)**")
        if st.button("Fetch & Append CHI Week to Workbook (Direct+Fallbacks)"):
            season_val, week_val = int(season_input), int(week_input)
            team_week = resolve_team_week_table(season_val, prefer_week=week_val)
            st.caption(f"Weekly table rows: {len(team_week)} (showing first few)")
            if not team_week.empty:
                st.dataframe(make_unique_columns(team_week.head(10)), use_container_width=True)

            chi = bears_row_for_week(team_week, season_val, week_val, NFL_TEAM)

            if chi.empty:
                # as a last-ditch, PBP per-team (quiet)
                pbp = _pbp_week_for_season(season_val)
                chi = _aggregate_team_week_from_pbp(pbp, NFL_TEAM, week_val)

            if chi.empty:
                st.warning(
                    "No CHI weekly row available yet from team-week, player-week, ESPN, or PBP. "
                    "This usually means the weekly assets haven’t been posted yet."
                )
            else:
                for col, val in (("season", season_val), ("week", week_val), ("team", NFL_TEAM)):
                    if col not in chi.columns:
                        chi[col] = val
                chi = make_unique_columns(chi)

                if not has_meaningful_numeric(chi):
                    st.info("Fetched a CHI row but it has no numeric values yet; not writing to workbook/CSVs.")
                else:
                    _off = append_to_sheet(EXCEL_PATH, SHEET_OFFENSE, chi, dedupe_on=["season", "week", "team"])
                    _def = append_to_sheet(EXCEL_PATH, SHEET_DEFENSE, chi, dedupe_on=["season", "week", "team"])
                    p1 = save_csv(chi, "Offense", season_val, week_val, NFL_TEAM)
                    p2 = save_csv(chi, "Defense", season_val, week_val, NFL_TEAM)
                    st.success(f"Workbook updated. CSVs written:\n• {os.path.basename(p1)}\n• {os.path.basename(p2)}")
                    st.dataframe(make_unique_columns(chi.head(50)), width="stretch")

        if st.button("Fetch & Write Opponent Week CSV (Direct+Fallbacks)"):
            season_val, week_val = int(season_input), int(week_input)
            team_week = resolve_team_week_table(season_val, prefer_week=week_val)
            opp = bears_row_for_week(team_week, season_val, week_val, opponent_input)

            if opp.empty:
                pbp = _pbp_week_for_season(season_val)
                opp = _aggregate_team_week_from_pbp(pbp, opponent_input, week_val)

            if opp.empty:
                st.warning("No opponent weekly row available yet.")
            else:
                for col, val in (("season", season_val), ("week", week_val), ("team", opponent_input)):
                    if col not in opp.columns:
                        opp[col] = val
                opp = make_unique_columns(opp)

                if not has_meaningful_numeric(opp):
                    st.info("Fetched an OPP row but it has no numeric values yet; not writing CSVs.")
                else:
                    p1 = save_csv(opp, "Offense", season_val, week_val, opponent_input)
                    p2 = save_csv(opp, "Defense", season_val, week_val, opponent_input)
                    st.success(f"Opponent CSVs written:\n• {os.path.basename(p1)}\n• {os.path.basename(p2)}")

        if st.button("Fetch & Write SnapCount CSVs (Auto)"):
            season_val, week_val = int(season_input), int(week_input)
            chi_sn, opp_sn = fetch_snap_counts_week_both(season_val, week_val, NFL_TEAM, opponent_input)
            wrote = False
            if not chi_sn.empty:
                p = save_csv(make_unique_columns(chi_sn), "SnapCounts", season_val, week_val, NFL_TEAM)
                st.success(f"Bears SnapCounts CSV: {os.path.basename(p)}")
                wrote = True
            if not opp_sn.empty:
                p = save_csv(make_unique_columns(opp_sn), "SnapCounts", season_val, week_val, opponent_input)
                st.success(f"Opponent SnapCounts CSV: {os.path.basename(p)}")
                wrote = True
            if not wrote:
                st.warning("No snap counts available to write yet.")

    with cc2:
        st.markdown("**Notes / Key Items**")
        key_notes = st.text_area("Quick Notes (saved under OpponentPreview)", height=120)
        if st.button("Save Notes to Opponent Preview"):
            if key_notes.strip():
                row = pd.DataFrame([{
                    "season": int(season_input),
                    "week": int(week_input),
                    "team": NFL_TEAM,
                    "opponent": opponent_input,
                    "Notes": key_notes.strip(),
                    "saved_at": datetime.now().isoformat(timespec="seconds")
                }])
                append_to_sheet(EXCEL_PATH, SHEET_OPP_PREVIEW, row,
                                dedupe_on=["season", "week", "team", "opponent"])
                st.success("Notes saved.")
            else:
                st.info("Nothing to save.")

# 2) Upload Weekly Data
with st.expander("2) Upload Weekly Data", expanded=True):
    st.caption("Upload Offense/Defense/Personnel/SnapCounts/Injuries CSVs. Rows are appended and deduped.")
    upc1, upc2 = st.columns(2)
    with upc1:
        st.markdown("**Offense CSV**")
        f_off = st.file_uploader("Upload Offense CSV", type=["csv"], key="up_off")
        if f_off is not None:
            try:
                df = pd.read_csv(f_off)
                for col, val in (("season", int(season_input)), ("week", int(week_input)), ("team", NFL_TEAM)):
                    if col not in df.columns: df[col] = val
                df = make_unique_columns(df.loc[:, ~df.columns.duplicated()])
                df_out = append_to_sheet(EXCEL_PATH, SHEET_OFFENSE, df, dedupe_on=["season", "week", "team"])
                st.success(f"Offense rows now: {len(df_out)}")
            except Exception as e:
                st.error(f"Upload failed: {e}")

        st.markdown("**Defense CSV**")
        f_def = st.file_uploader("Upload Defense CSV", type=["csv"], key="up_def")
        if f_def is not None:
            try:
                df = pd.read_csv(f_def)
                for col, val in (("season", int(season_input)), ("week", int(week_input)), ("team", NFL_TEAM)):
                    if col not in df.columns: df[col] = val
                df = make_unique_columns(df.loc[:, ~df.columns.duplicated()])
                df_out = append_to_sheet(EXCEL_PATH, SHEET_DEFENSE, df, dedupe_on=["season", "week", "team"])
                st.success(f"Defense rows now: {len(df_out)}")
            except Exception as e:
                st.error(f"Upload failed: {e}")

        st.markdown("**Injuries CSV**")
        f_inj = st.file_uploader("Upload Injuries CSV", type=["csv"], key="up_inj")
        if f_inj is not None:
            try:
                df = pd.read_csv(f_inj)
                for col, val in (("season", int(season_input)), ("week", int(week_input)), ("team", NFL_TEAM)):
                    if col not in df.columns: df[col] = val
                df = make_unique_columns(df.loc[:, ~df.columns.duplicated()])
                df_out = append_to_sheet(EXCEL_PATH, SHEET_INJURIES, df,
                                         dedupe_on=[c for c in ["season", "week", "team", "player"] if c in df.columns])
                st.success(f"Injuries rows now: {len(df_out)}")
            except Exception as e:
                st.error(f"Upload failed: {e}")

    with upc2:
        st.markdown("**Personnel CSV**")
        f_per = st.file_uploader("Upload Personnel CSV", type=["csv"], key="up_per")
        if f_per is not None:
            try:
                df = pd.read_csv(f_per)
                for col, val in (("season", int(season_input)), ("week", int(week_input)), ("team", NFL_TEAM)):
                    if col not in df.columns: df[col] = val
                df = make_unique_columns(df.loc[:, ~df.columns.duplicated()])
                df_out = append_to_sheet(EXCEL_PATH, SHEET_PERSONNEL, df, dedupe_on=["season", "week", "team"])
                st.success(f"Personnel rows now: {len(df_out)}")
            except Exception as e:
                st.error(f"Upload failed: {e}")

        st.markdown("**Snap Counts CSV (Manual)**")
        f_snap = st.file_uploader("Upload Snap Counts CSV", type=["csv"], key="up_snap")
        if f_snap is not None:
            try:
                df = pd.read_csv(f_snap)
                for col, val in (("season", int(season_input)), ("week", int(week_input)), ("team", NFL_TEAM)):
                    if col not in df.columns: df[col] = val
                df = make_unique_columns(df.loc[:, ~df.columns.duplicated()])
                dedupe_cols = ["season", "week", "team"]
                if "player" in df.columns:
                    dedupe_cols.append("player")
                df_out = append_to_sheet(EXCEL_PATH, SHEET_SNAP_COUNTS, df, dedupe_on=dedupe_cols)
                st.success(f"SnapCounts rows now: {len(df_out)}")
            except Exception as e:
                st.error(f"Upload failed: {e}")

# 3) NFL Averages (Auto & Manual)
with st.expander("3) NFL Averages (Auto & Manual)", expanded=True):
    st.markdown("Use the sidebar ‘Fetch NFL Data (Auto)’ to compute YTD league averages (weeks 1..W).")
    off_view = read_sheet(EXCEL_PATH, SHEET_YTD_NFL_OFF)
    if not off_view.empty:
        st.write("**NFL YTD Averages (Auto)**")
        st.dataframe(make_unique_columns(off_view), width="stretch")
    else:
        st.info("No auto NFL YTD averages saved yet.")

    st.markdown("---")
    st.write("**Selected Week NFL Average (live)**")
    wk_table = resolve_team_week_table(int(season_input), prefer_week=int(week_input))
    nfl_week_avg_row = nfl_week_average_from_table(wk_table, int(week_input))
    if nfl_week_avg_row.empty:
        st.caption("No league average for this week yet.")
    else:
        st.dataframe(make_unique_columns(nfl_week_avg_row), width="stretch")

    st.markdown("---")
    st.markdown("**Manual NFL Averages CSV** (optional). Columns will be prefixed to `NFL_Avg._` for styling.")
    f_nfl = st.file_uploader("Upload Manual NFL Averages CSV", type=["csv"], key="up_nflavg")
    if f_nfl is not None:
        try:
            df = pd.read_csv(f_nfl)
            df.columns = [c if c.startswith("NFL_Avg._") else f"NFL_Avg._{c}" for c in df.columns]
            df_out = append_to_sheet(EXCEL_PATH, SHEET_NFL_AVG_MANUAL, df, dedupe_on=None)
            st.success(f"Saved {len(df)} row(s) to NFL_Averages_Manual.")
            st.dataframe(make_unique_columns(df_out.tail(10)), width="stretch")
        except Exception as e:
            st.error(f"Upload failed: {e}")

# 4) Comparisons
with st.expander("4) DVOA Proxy, Color Codes & Comparisons", expanded=True):
    tabs = st.tabs(["Bears vs Opponent (This Week)", "Bears YTD vs NFL (Weeks 1..W)"])

    with tabs[0]:
        team_week = resolve_team_week_table(int(season_input), prefer_week=int(week_input))
        chi = bears_row_for_week(team_week, int(season_input), int(week_input), NFL_TEAM)
        opp = bears_row_for_week(team_week, int(season_input), int(week_input), opponent_input)
        if chi.empty and opp.empty:
            st.info("No weekly rows found yet. (Direct/agg/ESPN not posted; try again later.)")
        else:
            nums_chi = select_numeric(chi) if not chi.empty else []
            nums_opp = select_numeric(opp) if not opp.empty else []
            common = sorted(list(set(nums_chi).intersection(nums_opp)))
            if common:
                show_cols = ["season", "week", "team"] + common
                chi_show = chi[show_cols].add_suffix("_CHI") if not chi.empty else pd.DataFrame(columns=[c+"_CHI" for c in show_cols])
                opp_show = opp[show_cols].add_suffix("_OPP") if not opp.empty else pd.DataFrame(columns=[c+"_OPP" for c in show_cols])
                merged = pd.concat([chi_show.reset_index(drop=True), opp_show.reset_index(drop=True)], axis=1)
                merged = make_unique_columns(merged)
                st.dataframe(merged, width="stretch")
            else:
                st.dataframe(make_unique_columns(pd.concat([chi, opp], ignore_index=True)), width="stretch")

        st.markdown("**Bears vs NFL Average — This Week**")
        nfl_week_avg_row = nfl_week_average_from_table(team_week, int(week_input))
        if not chi.empty and not nfl_week_avg_row.empty:
            ids = [c for c in chi.columns if c in ("season", "week", "team")]
            nums = select_numeric(chi)
            preview = pd.concat([chi[ids + nums].reset_index(drop=True),
                                 nfl_week_avg_row.reset_index(drop=True)], axis=1)
            def _cell_color(val, base_col):
                nfl_col = f"NFL_Avg._{base_col}"
                if nfl_col not in preview.columns:
                    return ""
                ref = preview[nfl_col].iloc[0]
                if pd.isna(val) or pd.isna(ref):
                    return ""
                return "background-color: #e5ffe5" if val > ref else "background-color: #ffe5e5"
            def _apply(row):
                styles = []
                for c in preview.columns:
                    if c in ids or c.startswith("NFL_Avg._"):
                        styles.append("")
                    else:
                        styles.append(_cell_color(row[c], c))
                return styles
            st.dataframe(make_unique_columns(preview).style.apply(_apply, axis=1), width="stretch")
        else:
            st.caption("Need CHI row and NFL week average to color-code.")

        st.markdown("**Snap Counts (This Week — Bears only)**")
        chi_sn, _ = fetch_snap_counts_week_both(int(season_input), int(week_input), NFL_TEAM, opponent_input)
        if chi_sn.empty:
            st.caption("No CHI snap counts this week yet.")
        else:
            st.dataframe(make_unique_columns(chi_sn.head(50)), width="stretch")

    with tabs[1]:
        team_week = resolve_team_week_table(int(season_input), prefer_week=int(week_input))
        bears_ytd = fetch_bears_averages_YTD(int(season_input), int(week_input), NFL_TEAM, team_week)
        nfl_ytd = fetch_nfl_averages_YTD(int(season_input), int(week_input), team_week)
        if bears_ytd.empty or nfl_ytd.empty:
            st.info("Need both Bears YTD and NFL YTD (try ‘Fetch NFL Data (Auto)’ after weekly posts).")
        else:
            st.dataframe(make_unique_columns(pd.concat([bears_ytd.reset_index(drop=True),
                                                        nfl_ytd.reset_index(drop=True)], axis=1)),
                         width="stretch")

# 5) Opponent Preview & Strategy Notes
with st.expander("5) Opponent Preview & Strategy Notes", expanded=True):
    st.caption("Notes and predictions you can save and review.")
    opp = read_sheet(EXCEL_PATH, SHEET_OPP_PREVIEW)
    if not opp.empty:
        st.write("**Opponent Preview (Recent)**")
        st.dataframe(make_unique_columns(opp.sort_values(opp.columns[0]).tail(25)), width="stretch")
    else:
        st.info("No opponent preview entries yet.")
    colA, colB = st.columns([2, 1])
    with colA:
        rationale = st.text_area("Rationale", height=120, key="pred_rationale")
    with colB:
        predicted_winner = st.selectbox("Predicted Winner", options=[NFL_TEAM, opponent_input], index=0)
        confidence = st.slider("Confidence", min_value=0, max_value=100, value=60, step=5)
    if st.button("Save Weekly Prediction"):
        row = pd.DataFrame([{
            "season": int(season_input),
            "week": int(week_input),
            "team": NFL_TEAM,
            "opponent": opponent_input,
            "Predicted_Winner": predicted_winner,
            "Confidence": confidence,
            "Rationale": rationale.strip(),
            "saved_at": datetime.now().isoformat(timespec="seconds")
        }])
        df_out = append_to_sheet(EXCEL_PATH, SHEET_PREDICTIONS, row,
                                 dedupe_on=["season", "week", "team", "opponent"])
        st.success("Prediction saved.")
        st.dataframe(make_unique_columns(df_out.tail(10)), width="stretch")

# 6) Exports & Downloads
with st.expander("6) Exports & Downloads", expanded=True):
    st.caption("Create week-only Excel, final PDF, and download the full workbook.")

    st.markdown("**Weekly Excel (This Week)**")
    if st.button("Create Weekly Excel (Off/Def/Personnel/Snaps/Injuries)"):
        season_val = int(season_input); week_val = int(week_input)
        def fw(df):
            if df.empty: return df
            if 'season' in df.columns: df = df[df['season'] == season_val]
            if 'week' in df.columns: df = df[df['week'] == week_val]
            return df
        off = fw(read_sheet(EXCEL_PATH, SHEET_OFFENSE))
        deff = fw(read_sheet(EXCEL_PATH, SHEET_DEFENSE))
        per = fw(read_sheet(EXCEL_PATH, SHEET_PERSONNEL))
        snaps = fw(read_sheet(EXCEL_PATH, SHEET_SNAP_COUNTS))
        inj = fw(read_sheet(EXCEL_PATH, SHEET_INJURIES))
        team_week = resolve_team_week_table(season_val, prefer_week=week_val)
        chi = bears_row_for_week(team_week, season_val, week_val, NFL_TEAM)
        opp = bears_row_for_week(team_week, season_val, week_val, opponent_input)
        nfl_wk = nfl_week_average_from_table(team_week, week_val)
        bio = io.BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as xlw:
            if not off.empty:   make_unique_columns(off).to_excel(xlw, index=False, sheet_name="Offense")
            if not deff.empty:  make_unique_columns(deff).to_excel(xlw, index=False, sheet_name="Defense")
            if not per.empty:   make_unique_columns(per).to_excel(xlw, index=False, sheet_name="Personnel")
            if not snaps.empty: make_unique_columns(snaps).to_excel(xlw, index=False, sheet_name="SnapCounts")
            if not inj.empty:   make_unique_columns(inj).to_excel(xlw, index=False, sheet_name="Injuries")
            if not chi.empty:   make_unique_columns(chi).to_excel(xlw, index=False, sheet_name="CHI_Week")
            if not opp.empty:   make_unique_columns(opp).to_excel(xlw, index=False, sheet_name=f"{opponent_input}_Week")
            if not nfl_wk.empty: make_unique_columns(nfl_wk).to_excel(xlw, index=False, sheet_name="NFL_Week_Avg")
        bio.seek(0)
        st.download_button(
            label=f"Download Weekly Excel — {int(season_input)} W{int(week_input):02d}",
            data=bio.getvalue(),
            file_name=f"bears_week_{int(season_input)}_W{int(week_input):02d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )

    st.markdown("---")
    if st.button("Export Final PDF (This Week)"):
        try:
            team_week = resolve_team_week_table(int(season_input), prefer_week=int(week_input))
            lines = build_pdf_summary_lines(int(season_input), int(week_input), opponent_input, team_week)
            pdf = export_week_pdf(int(season_input), int(week_input), opponent_input, lines)
            if not pdf:
                st.error("FPDF not available — ensure fpdf is in requirements.")
            else:
                out_path = os.path.join(EXPORTS_DIR, f"{int(season_input)}_W{int(week_input):02d}_{opponent_input}_Final.pdf")
                with open(out_path, "wb") as f:
                    f.write(pdf)
                st.success(f"Final PDF created: {os.path.basename(out_path)}")
                st.download_button("Download Final PDF (Just Created)", data=pdf,
                                   file_name=os.path.basename(out_path), mime="application/pdf", type="primary")
        except Exception as e:
            st.error(f"PDF export failed: {e}")

    st.markdown("---")
    if os.path.exists(EXCEL_PATH):
        with open(EXCEL_PATH, "rb") as f:
            st.download_button(
                label="Download Excel (All Data)",
                data=f.read(),
                file_name="bears_weekly_analytics.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    else:
        st.info("Excel not found yet; upload or fetch to create it.")

    st.markdown("**Workbook Peek**")
    cols = st.columns(3)
    try:
        with cols[0]:
            st.write("Offense")
            off = read_sheet(EXCEL_PATH, SHEET_OFFENSE)
            st.dataframe(make_unique_columns(off.tail(10)), width="stretch") if not off.empty else st.caption("—")
        with cols[1]:
            st.write("Defense")
            deff = read_sheet(EXCEL_PATH, SHEET_DEFENSE)
            st.dataframe(make_unique_columns(deff.tail(10)), width="stretch") if not deff.empty else st.caption("—")
        with cols[2]:
            st.write("SnapCounts")
            snaps = read_sheet(EXCEL_PATH, SHEET_SNAP_COUNTS)
            st.dataframe(make_unique_columns(snaps.tail(10)), width="stretch") if not snaps.empty else st.caption("—")
        st.write("Injuries")
        inj = read_sheet(EXCEL_PATH, SHEET_INJURIES)
        st.dataframe(make_unique_columns(inj.tail(10)), width="stretch") if not inj.empty else st.caption("—")
    except Exception as e:
        st.error(f"Peek failed: {e}")
